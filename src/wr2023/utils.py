import statistics
import zarr 
import pathlib
import csv
from glob import glob
import scipy as sp 
import numpy as np 
from openpyxl import load_workbook

def selection_criteria():
    # return [50]
    return sorted( 
        [83, 86, 98, 101, 104, 144,
        149, 159, 164, 176, 216, 240,
        243, 247, 262, 269, 317, 318,
        319, 50, 52, 56, 61, 66, 70,
        74, 77, 80, 219, 224, 227,
        230, 233, 250, 253, 255,
        260, 320, 321, 322, 323, 324 ]
    )

def report_model_metrics(DST_PATH):
    root = zarr.open(DST_PATH, mode='r')

    HEADER = ['id', '# nbeats', '# nbad', 'cc', 're'] # 'rdms', 'rmse', 'mae']
    file = pathlib.Path(__file__).parent / 'reports' / 'model_metrics.csv'

    with open(file, 'w', newline='') as csvfile:
        filewriter = csv.writer(csvfile, delimiter='\t')
        filewriter.writerow(HEADER)

        for n in selection_criteria():
            gt_group = root[f'{n}/ground_truth/']
            mv_group = root[f'{n}/model_validation/']

            metrics = mv_group['metrics'][:]
            gt_bad_beats = gt_group['bad_beats'][:]
            # tr_bad_beats = root[f'{n}/transform_validation/bad_beats'][:]

            nbeats = gt_group['beat_matrix'].shape[0] # amount of beats
            nbad = gt_bad_beats.size # amount of bad beats
            # neff = nbeats - nbad # amount of beats for model validations
            # nefftr = np.union1d(gt_bad_beats, tr_bad_beats).size - nbad # extra amount of beats for model validations

            cc = f'{np.nanmean(metrics[:,0]):.4f}({np.nanstd(metrics[:,0]):.4f})'
            re = f'{np.nanmean(metrics[:,4]):.4f}({np.nanstd(metrics[:,4]):.4f})'
            # rdms = f'{np.nanmean(metrics[:,1]):.4f}({np.nanstd(metrics[:,1]):.4f})'
            # rmse = f'{np.nanmean(metrics[:,2]):.4f}({np.nanstd(metrics[:,2]):.4f})'
            # mae = f'{np.nanmean(metrics[:,3]):.4f}({np.nanstd(metrics[:,3]):.4f})'

            filewriter.writerow([f'{n:03}', f'{nbeats:04}', f'{nbad:04}', cc, re ]) # rdms, rmse, mae, re])

def report_transform_metrics(DST_PATH):
    root = zarr.open(DST_PATH, mode='r')

    HEADER = ['id', '# nbeats', '# nbad', 'P', 'PR', 'RS', 'QT', 'P', 'R', 'S', 'T']
    file = pathlib.Path(__file__).parent / 'reports' / 'transform_metrics.csv'
    with open(file, 'w', newline='') as csvfile:

        filewriter = csv.writer(csvfile, delimiter='\t')
        filewriter.writerow(HEADER)

        for n in selection_criteria():
            gt_group = root[f'{n}/ground_truth']
            tr_group = root[f'{n}/transform_validation']
            bad_beats = np.union1d(gt_group['bad_beats'][:], tr_group['bad_beats'][:])

            nbeats = gt_group['beat_matrix'].shape[0] # amount of beats
            nbad = bad_beats.size # amount of bad beats

            gt_mea = np.delete(np.delete(gt_group['measurements'][:], bad_beats, axis=0), 5, axis=1)
            tr_mea = np.delete(np.delete(tr_group['measurements'][:], bad_beats, axis=0), 5, axis=1)

            # if runtime warning raised implies null std (measurements are equal in all beats)
            rho = np.corrcoef(gt_mea, y=tr_mea, rowvar=False).diagonal(8) 
            rho = np.round(rho, decimals=4)

            filewriter.writerow([f'{n:03}', f'{nbeats:04}', f'{nbad:04}', *rho.tolist() ]) # rdms, rmse, mae, re])

def report_beat_criteria(DST_PATH, group_name='ground_truth'):
    root = zarr.open(DST_PATH, mode='r')
    
    HEADER = ['id', 'invalid fiducial mask', 'lesser cc mask', 'out of window mask', 'lost [%]']
    file = pathlib.Path(__file__).parent / 'reports' / f'beat_criteria_of_{group_name}.csv'

    with open(file, 'w', newline='') as csvfile:
        filewriter = csv.writer(csvfile, delimiter='\t')
        filewriter.writerow(HEADER)

        for n in selection_criteria():
            group = root[f'{n}/{group_name}/']
            bad_beats = group['bad_beats']
            n_beats = group['beat_matrix'].shape[0]
            att = bad_beats.attrs.items()
            row = [None]*len(HEADER)
            row[0] = n
            for k, v in att:
                ix = HEADER.index(k)
                row[ix] = v
            row[-1] = f'{100 * ( row[1] + row[2] + row[3] ) / n_beats :.1f} ({n_beats})'
            filewriter.writerow(row)

def import_matfiles(SRC_PATH):
    females_paths = glob('E:\paper WR model\hembras lead II\*.mat')
    males_paths = glob('E:\paper WR model\machos lead II\*.mat')
    paths = females_paths + males_paths
    for f in paths:
        n = pathlib.Path(f).stem.split('_')[0]
        fmat = sp.io.loadmat(f, simplify_cells=True)['data']
        sp.io.savemat(SRC_PATH + n + '.mat', {'DII':fmat})

def export_matfiles(DST_PATH, group_name='ground_truth'):
    root = zarr.open(DST_PATH, mode='r')
    db = {}
    for n in selection_criteria():
        group = root[f'{n}/{group_name}']
        ecg = group[f'ecg'][:]
        try:
            n_beats = group[f'{group_name}/beat_matrix'].shape[0]
        except:
            n_beats = -1
        db[f'Reg{n}'] = {'ecg': ecg, 'beats':n_beats}
    sp.io.savemat(f'{group_name}.mat', db)

def stats(XLSX_PATH):
    ir = InfoReader(XLSX_PATH)

    males = [v for v in ir.get_males() if v in selection_criteria()]
    females = [v for v in ir.get_females() if v in selection_criteria()]

    m = ir.get(males)
    print(f'Machos {len(m)}')
    print(f"Peso = {ir.mean(m, 'peso'):.1f}({ir.std(m, 'peso'):.1f})")
    print(f"Edad = {ir.mean(m, 'edad'):.1f}({ir.std(m, 'edad'):.1f})")

    m = ir.get(females)
    print(f'Hembras {len(m)}')
    print(f"Peso = {ir.mean(m, 'peso'):.1f}({ir.std(m, 'peso'):.1f})")
    print(f"Edad = {ir.mean(m, 'edad'):.1f}({ir.std(m, 'edad'):.1f})")


class InfoReader:
    def __init__(self, path) -> None:
        sheet = load_workbook(filename = path).active
        self._data = {}
        for row in sheet.iter_rows(min_row=4, max_col=16, max_row=220):
            n = row[0].value
            self._data[n] = {
                'rata': row[1].value,
                'fecha de nacimiento': row[2].value,
                'fecha de registro': row[3].value,
                'fecha de castracion': row[4].value,
                'peso': row[5].value,
                'edad': (row[3].value - row[2].value).total_seconds() / 60 / 60 / 24 / 7,
                'sexo': self.set_sex(row),
                'estado': self.set_state(row)
            }

    def set_state(self, row): 
        if row[6].value == 1:
            return 'C1'
        if row[7].value == 1:
            return 'C2'
        if row[8].value == 1:
            return 'A1'
        if row[9].value == 1:
            return 'A2'
        if row[10].value == 1:
            return 'A3'
        if row[11].value == 1:
            return 'P1'
        if row[12].value == 1:
            return 'P2'
        raise ValueError(f'Undefined state in register {row[0].value}')
    
    def set_sex(self, row):
        if row[14].value == 1:
            return 'F'
        if row[15].value == 1:
            return 'M'
        raise ValueError(f'Undefined sex in register {row[0].value}')
    
    def get_males(self):
        return [ k for k, v in self._data.items() if v['sexo'] == 'M'] 

    def get_females(self):
        return [ k for k, v in self._data.items() if v['sexo'] == 'F'] 
    
    def get(self, lst):
        return { k:self._data[k] for k in lst} 
    
    def mean(self, data, key):
        return statistics.mean(map(lambda r: r[key], data.values()))
    
    def std(self, data, key):
        return statistics.stdev(map(lambda r: r[key], data.values()))